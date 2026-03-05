# services/excel_service.py
# Generate an in-memory Excel workbook from the stored expenses

import io
from typing import List

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.models.schemas import Expense


# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_BG   = "1E3A5F"   # dark navy
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "EBF2FA"   # light blue tint
SUMMARY_BG  = "F0F7EE"   # light green tint
ACCENT      = "2563EB"   # blue accent for totals row

THIN = Side(style="thin", color="BFCFE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _currency(ws, cell_ref, value: float):
    """Write a float as a currency-formatted cell."""
    cell = ws[cell_ref]
    cell.value = value
    cell.number_format = '#,##0.00'
    return cell


def generate_excel(expenses: List[Expense]) -> bytes:
    """
    Build an Excel workbook with:
    - 'Expenses' sheet – full expense table
    - 'Summary'  sheet – high-level aggregates
    Returns raw bytes ready to stream to the client.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1 – Expenses ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Expenses"

    columns = ["Bill Name", "Vendor", "Category", "Expense Type",
               "Amount ($)", "GST ($)", "Date"]
    col_widths = [28, 24, 18, 16, 14, 12, 16]

    # Header row
    for col_idx, (header, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    total_amount = 0.0
    total_gst    = 0.0

    for row_idx, exp in enumerate(expenses, start=2):
        row_data = [
            exp.bill_name,
            exp.vendor,
            exp.category,
            exp.expense_type,
            exp.amount,
            exp.gst,
            exp.date or "",
        ]
        fill = PatternFill("solid", fgColor=ALT_ROW_BG) if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = BORDER
            if fill:
                cell.fill = fill
            if col_idx in (5, 6):          # currency columns
                cell.number_format = '#,##0.00'

        total_amount += exp.amount
        total_gst    += exp.gst

    # Totals row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True, size=11)
    for col_idx in (5, 6):
        cell = ws.cell(
            row=total_row, column=col_idx,
            value=total_amount if col_idx == 5 else total_gst,
        )
        cell.font         = Font(bold=True, color="FFFFFF", size=11)
        cell.fill         = PatternFill("solid", fgColor=ACCENT)
        cell.number_format = '#,##0.00'
        cell.border        = BORDER

    ws.freeze_panes = "A2"

    # ── Sheet 2 – Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    business = sum(e.amount for e in expenses if e.expense_type == "Business")
    personal  = sum(e.amount for e in expenses if e.expense_type == "Personal")

    summary_data = [
        ("Metric",              "Value"),
        ("Total Expenses",      total_amount),
        ("Total GST",           total_gst),
        ("Business Expenses",   business),
        ("Personal Expenses",   personal),
        ("Number of Bills",     len(expenses)),
    ]

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 18

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        lc = ws2.cell(row=r_idx, column=1, value=label)
        vc = ws2.cell(row=r_idx, column=2, value=value)

        if r_idx == 1:
            lc.font = Font(bold=True, color=HEADER_FG)
            vc.font = Font(bold=True, color=HEADER_FG)
            lc.fill = PatternFill("solid", fgColor=HEADER_BG)
            vc.fill = PatternFill("solid", fgColor=HEADER_BG)
        else:
            lc.fill = PatternFill("solid", fgColor=SUMMARY_BG)
            if isinstance(value, float):
                vc.number_format = '#,##0.00'

        for cell in (lc, vc):
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = BORDER

        ws2.row_dimensions[r_idx].height = 22

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
